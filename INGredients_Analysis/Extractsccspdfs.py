import openpyxl
import requests
import hashlib
import os
import json
import re
from urllib.parse import urlparse
import time

# --- 配置 ---

# 您的Excel文件路径
EXCEL_FILE_PATH = 'Cosing数据库XLS2.XLSX'
# 您要保存PDF的文件夹名称
DOWNLOAD_FOLDER = 'sccspdfs'
# PDF链接所在的列 (H列是第8列)
SOURCE_COLUMN_INDEX = 8  # H
# 保存JSON结果的列 (N列是第14列)
TARGET_COLUMN_INDEX = 14 # N

# --- ---

def find_pdf_urls(text):
    """
    使用正则表达式从文本中查找所有以 .pdf 结尾的 URL。
    """
    if not isinstance(text, str):
        return []
    
    # 这个正则表达式会查找 http:// 或 https:// 开头，
    # 后面跟着一系列非空白、非引号的字符，并以 .pdf 结尾
    # re.IGNORECASE 使得 .pdf 不区分大小写
    regex = r'https?://[^\s;"\'<>]+?\.pdf'
    urls = re.findall(regex, text, re.IGNORECASE)
    
    # 去除重复的链接（同时保持顺序）
    unique_urls = list(dict.fromkeys(urls))
    return unique_urls

def download_and_hash_pdf(url, folder):
    """
    下载指定URL的PDF文件，计算其SHA256哈希值，并以哈希值命名保存。
    返回一个包含处理结果的字典。
    """
    try:
        # 提取URL中的原始文件名
        original_filename = os.path.basename(urlparse(url).path)
        if not original_filename:
            original_filename = "unknown.pdf"

        # 发起网络请求
        # 设置超时以防止程序卡在无响应的链接
        response = requests.get(url, timeout=20)
        # 如果下载失败（如 404 Not Found），则引发异常
        response.raise_for_status()
        
        pdf_content = response.content
        
        # 计算文件内容的SHA256哈希值
        file_hash = hashlib.sha256(pdf_content).hexdigest()
        
        # 构建保存路径
        hash_filename = f"{file_hash}.pdf"
        save_path = os.path.join(folder, hash_filename)
        
        # 写入文件
        with open(save_path, 'wb') as f:
            f.write(pdf_content)
            
        # 返回成功结果
        return {
            "status": "success",
            "original_url": url,
            "original_filename": original_filename,
            "hash_filename": hash_filename
        }
        
    except requests.exceptions.RequestException as e:
        # 处理所有网络或HTTP错误
        print(f"   [失败] 下载链接 {url} 时出错: {e}")
        return {
            "status": "error",
            "original_url": url,
            "message": str(e)
        }
    except Exception as e:
        # 处理其他未知错误
        print(f"   [失败] 处理链接 {url} 时发生未知错误: {e}")
        return {
            "status": "error",
            "original_url": url,
            "message": f"An unexpected error occurred: {str(e)}"
        }

def process_excel(file_path, download_folder):
    """
    主处理函数：遍历Excel，根据H列下载PDF，并将结果写入N列。
    """
    
    print(f"开始处理Excel文件: {file_path}")
    
    # 1. 确保下载文件夹存在
    os.makedirs(download_folder, exist_ok=True)
    print(f"PDF将保存到: {download_folder}")

    try:
        # 2. 加载工作簿和活动工作表
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        print(f"成功加载工作表: {sheet.title} (共 {sheet.max_row} 行)")
    except FileNotFoundError:
        print(f"错误：找不到Excel文件，请检查路径: {file_path}")
        return
    except Exception as e:
        print(f"打开Excel文件时出错: {e}")
        return

    processed_count = 0
    skipped_count = 0
    
    # 3. 遍历所有行 (假设第1行是标题，从第2行开始)
    # iter_rows 返回的是元组，索引从0开始
    # 所以 H列(8) 对应索引 7, N列(14) 对应索引 13
    for row in sheet.iter_rows(min_row=2):
        
        # 使用 .cell(row=, column=) 会更稳定
        current_row_index = row[0].row
        source_cell = sheet.cell(row=current_row_index, column=SOURCE_COLUMN_INDEX)
        target_cell = sheet.cell(row=current_row_index, column=TARGET_COLUMN_INDEX)
        
        source_data = source_cell.value
        target_data = target_cell.value
        
        # 4. 检查处理条件：H列有内容，且N列为空
        if source_data and not target_data:
            print(f"\n[处理中] 第 {current_row_index} 行...")
            
            # 5. 从H列单元格中提取所有PDF链接
            urls = find_pdf_urls(source_data)
            
            if not urls:
                print("   - H列中未找到有效的 .pdf 链接。")
                target_cell.value = json.dumps([{"status": "error", "message": "No valid .pdf links found in source cell."}])
                continue

            print(f"   - 找到 {len(urls)} 个链接: {urls}")
            
            # 存储这一行所有链接的处理结果
            results_list = []
            
            # 6. 循环下载和处理找到的每个链接
            for i, url in enumerate(urls, 1):
                print(f"   - 正在处理第 {i}/{len(urls)} 个链接: {url[:70]}...")
                result = download_and_hash_pdf(url, download_folder)
                results_list.append(result)
                # 稍微暂停一下，避免请求过于频繁
                time.sleep(0.5)
            
            # 7. 将所有结果的列表转换为JSON字符串，写入N列
            # ensure_ascii=False 确保中文等字符正确显示
            # indent=2 使JSON格式化，更易读
            json_output = json.dumps(results_list, ensure_ascii=False, indent=2)
            target_cell.value = json_output
            print(f"   - 已将JSON结果写入 N{current_row_index}。")
            processed_count += 1
            
        elif target_data:
            skipped_count += 1
            # (静默跳过，避免过多日志)
            # print(f"[跳过] 第 {current_row_index} 行 (N列已有数据)。")
        else:
            skipped_count += 1
            # (静默跳过 H列为空的行)

    print(f"\n--------------------\n处理完毕。")
    print(f"总共处理并更新了 {processed_count} 行。")
    print(f"跳过了 {skipped_count} 行 (H列为空或N列已有数据)。")
    
    # 8. 保存修改后的Excel文件
    # 为了安全起见，我们保存为一个新文件，避免覆盖原文件
    save_path = file_path.replace('.xlsx', '_processed.xlsx')
    try:
        workbook.save(save_path)
        workbook.close()
        print(f"成功保存文件到: {save_path}")
    except PermissionError:
        print(f"\n[错误] 保存文件失败！")
        print(f"请确保您已关闭Excel文件 '{save_path}' (如果它已打开)，然后重试。")
    except Exception as e:
        print(f"\n[错误] 保存文件时发生未知错误: {e}")

# --- 脚本入口 ---
if __name__ == "__main__":
    # ！！！请确保在这里修改 EXCEL_FILE_PATH 为您的文件名 ！！！
    # EXCEL_FILE_PATH = '您的文件名.xlsx' 
    
    if EXCEL_FILE_PATH == 'your_excel_file.xlsx':
        print("="*50)
        print("警告：请先在脚本中修改 'EXCEL_FILE_PATH' 变量，")
        print("      将其指向您要处理的实际 Excel 文件路径。")
        print("="*50)
    else:
        process_excel(EXCEL_FILE_PATH, DOWNLOAD_FOLDER)